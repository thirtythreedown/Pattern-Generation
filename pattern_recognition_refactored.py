##Working with Excel Spreadsheets, from Automate the boring stuff with Python - https://automatetheboringstuff.com/2e/chapter13/
##Tutorial for OpenPyxl library - https://openpyxl.readthedocs.io/en/stable/tutorial.html

##DEPENDENCIES
##random library - built-in Python library
##openpyxl - pip install openpyxl
##LibreOffice, free tool for reading/creating .xlsx spreadsheets and workbooks - https://www.libreoffice.org/
 
##DONE
##Proof of concept
##Adding redundancy detection (Compare with list, if not in list, append?)

##TO DO
##Appending generated names to list of previously generated names

import random
import openpyxl

##Loading openpyxl module for working with workbooks/spreadsheets
workbook = openpyxl.load_workbook('pattern_recognition.xlsx')
##Loading pattern_recognition.xlsx workbook in wb variable using openpyxl.loard_workbook()

sheet = workbook['Sheet1']
writingsheet = workbook['Sheet2']
##Loading specific spreadsheet from workbook by name using ['sheetname'] method, storing it into sheet variable

##---PROBING SHEET SIZE STARTS HERE---
##print("How big is the sheet?")
##lastCol = sheet.get_highest_column() is deprecated! Use line below instead.
lastCol = sheet.max_column
##Looking at the biggest column number in the spreadsheet and storing in lastCol variable
lastRow = sheet.max_row
##print("The sheet is " + str(lastCol) + " columns wide")
##print("The sheet is " + str(lastRow) + " rows tall")
##---PROBING SHEET SIZE ENDS HERE---

colA = sheet['A']
##Loading the contents of column A in variable colA
##for cell in colA:
##    ##For each cell in ColA...
##    print(cell.coordinate, cell.value)
##    #Printing the contents of the cell
##print('---END OF COLUMN A---')

colB = sheet['B']
##Loading the contents of column A in variable colA
##for cell in colB:
##    ##For each cell in ColA...
##    print(cell.coordinate, cell.value)
##    #Printing the contents of the cell
##print('---END OF COLUMN B---')

colC = sheet['C']
##Loading the contents of column A in variable colA
##for cell in colC:
##    ##For each cell in ColA...
##    print(cell.coordinate, cell.value)
##    #Printing the contents of the cell
##print('---END OF COLUMN C---')

saved_projects = writingsheet['A']
##Loading sheet 2's column A into program as projects_list
##Need to load the .value() values of the cells into a list

counter = 0
##Setting counter to 0

generated_names=[]

def randomizer(sheet, column):
    """Defining a randomizing module"""
    randomized = sheet[column+str(random.randint(1, lastRow))].value
    ##print(randomized)
    return randomized

def list_cleaner(input_list, comparator_list):
    """Compares generated_names with saved_projects and removes duplicates"""  
    for item in input_list:
        print("Processing item " + item + " from list1")
        if item not in comparator_list:
            print("That's the stuff!")
        else:
            print("That's a duplicate!")
            input_list.remove(item)
    return input_list
   

print(saved_projects)

while counter < 5:
    ##Accessing a random cell from each column
    random_1 = randomizer(sheet, 'A')
    random_2 = randomizer(sheet, 'B')
    random_3 = randomizer(sheet, 'C')
    project_name = random_1 + ' ' + random_2 + ' ' + random_3
    print(project_name)
    generated_names.append(project_name)
    counter += 1
print(generated_names)
cleaned_list = list_cleaner(generated_names, saved_projects)

for value in cleaned_list:
    print(value)



##workbook.save('pattern_recognition.xlsx')
##    random_1 = sheet['A'+str(random.randint(1, lastRow))].value
##    random_2 = sheet['B'+str(random.randint(1, lastRow))].value
##    random_3 = sheet['C'+str(random.randint(1, lastRow))].value

##    project_name = random_1 + ' ' + random_2 + ' ' + random_3.lower()
##    print(project_name)
##    counter += 1
