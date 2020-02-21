#! python3

##Working with Excel Spreadsheets, from Automate the boring stuff with Python - https://automatetheboringstuff.com/2e/chapter13/
##Tutorial for OpenPyxl library - https://openpyxl.readthedocs.io/en/stable/tutorial.html

##DEPENDENCIES
##random library - built-in Python library
##openpyxl - pip install openpyxl
##LibreOffice, free tool for reading/creating .xlsx spreadsheets and workbooks - https://www.libreoffice.org/

import random
##Importing random module for random goodness
import openpyxl
##Importing openpyxl module for workbooks and spreadsheets goodness
workbook = openpyxl.load_workbook('pattern_generation.xlsx')
##Loading pattern_recognition.xlsx workbook in workbook variable

loading_sheet = workbook['name_elements']
writing_sheet = workbook['project_names']
##Loading spreadsheets from workbook by name, storing them into separate variables

lastCol = loading_sheet.max_column
##Looking at the biggest column number in the spreadsheet and storing in lastCol variable
lastRow = loading_sheet.max_row
##Looking at the biggest row number in the spreadsheet and storing in lastRow variable

colA = loading_sheet['A']
##Loading the contents of column A in variable colA

colB = loading_sheet['B']
##Loading the contents of column A in variable colB

colC = loading_sheet['C']
##Loading the contents of column A in variable colC

saved_projects = writing_sheet['A']
##Loading sheet 2's column A into program in saved_projects variable

counter = 0
##Setting counter to 0

generated_names=[]
##Creating empty generated_names list

def randomizer(sheet, column):
    """Defining a randomizing module"""
    randomized = sheet[column + str(random.randint(1, lastRow))].value
    ##Storing a random value from a column into the randomized variable
    return randomized
    ##Returning the randomized variable

def list_cleaner(input_list, comparator_list):
    """Compares generated_names with saved_projects and removes duplicates"""  
    for item in input_list:
    ##Looping through the items in input_list
        if item in comparator_list:
            input_list.remove(item)
        ##Removing item from the input_list if it already exists in comparator_list
    return input_list
    ##Returning the cleaned version of input_list

def list_saver(cleaned_list, writing_sheet, write_row):
    """Appends the cleaned-up list of the names to the existing list in the second sheet of the workbook"""  
    for item in cleaned_list:
    ##Looping through the contents of cleaned_list
        writing_sheet.cell(write_row, column = 1).value = item
        ##Updating the cells in the spreadsheet with the item
        write_row = write_row + 1
        ##Moving to the next row
    print("Done appending data!")

##print(saved_projects)

print("Generating names!")

while counter < 1000:
    ##Looping through the counter 1000-ish times
    random_1 = randomizer(loading_sheet, 'A')
    random_2 = randomizer(loading_sheet, 'B')
    random_3 = randomizer(loading_sheet, 'C')
    ##Getting random values from columns A, B and C
    project_name = random_1 + ' ' + random_2 + ' ' + random_3
    ##Putting together project_name from random values
    generated_names.append(project_name)
    ##Adding generated names to generated_names list
    counter += 1
    ##Updating counter

##print(generated_names)
##Uncomment above if you want to see the whole list of generated_names

print("Cleaning up the list of generated names!")
cleaned_list = list_cleaner(generated_names, saved_projects)
##Cleaning up the generated_names list

##print(cleaned_list)
##Uncomment above if you want to see the cleaned list names

write_row = writing_sheet.max_row + 1
##Getting the last row from the writing_sheet spreadsheet

print("Saving the list of generated names!")
list_saver(cleaned_list, writing_sheet, write_row)
 
print("Saving the updated workbook!")
workbook.save('pattern_generation.xlsx')
##Saving the updated workbook
print('All done!')
