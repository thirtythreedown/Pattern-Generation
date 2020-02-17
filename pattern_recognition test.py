##Working with Excel Spreadsheets, from Automate the boring stuff with Python - https://automatetheboringstuff.com/2e/chapter13/
##Tutorial for OpenPyxl library at https://openpyxl.readthedocs.io/en/stable/tutorial.html

##DEPENDENCIES
##random library - built-in Python library
##openpyxl - pip install openpyxl
##LibreOffice, free tool for reading/creating .xlsx spreadsheets and workbooks - https://www.libreoffice.org/
 
##DONE
##Proof of concept

##TO DO
##Adding redundancy detection (Compare with list, if not in list, append?)

import random
import openpyxl

##Loading openpyxl module for working with workbooks/spreadsheets
workbook = openpyxl.load_workbook('pattern_recognition.xlsx')
##Loading pattern_recognition.xlsx workbook in wb variable using openpyxl.loard_workbook()

sheet = workbook['Sheet1']
##Loading specific spreadsheet from workbook by name using ['sheetname'] method, storing it into sheet variable
##print(sheet)
##Printing content of worksheet object stored in sheet variable

##print("Printing the value of A1 cell")
##print(sheet['A1'].value)
##Printing the value of cell A1 in Sheet1. Yes, it's a list kinda thing.    

##---PROBING SHEET SIZE STARTS HERE---
##print("How big is the sheet?")
##lastCol = sheet.get_highest_column() is deprecated! Use line below instead.
lastCol = sheet.max_column
##Looking at the biggest column number in the spreadsheet and storing in lastCol variable
lastRow = sheet.max_row
##print("The sheet is " + str(lastCol) + " columns wide")
##print("The sheet is " + str(lastRow) + " rows tall")
##---PROBING SHEET SIZE ENDS HERE---

colA = sheet['A']
##Loading the contents of column A in variable colA
##for cell in colA:
##    ##For each cell in ColA...
##    print(cell.coordinate, cell.value)
##    #Printing the contents of the cell
##print('---END OF COLUMN A---')

colB = sheet['B']
##Loading the contents of column A in variable colA
##for cell in colB:
##    ##For each cell in ColA...
##    print(cell.coordinate, cell.value)
##    #Printing the contents of the cell
##print('---END OF COLUMN B---')

colC = sheet['C']
##Loading the contents of column A in variable colA
##for cell in colC:
##    ##For each cell in ColA...
##    print(cell.coordinate, cell.value)
##    #Printing the contents of the cell
##print('---END OF COLUMN C---')

counter = 0
##Setting counter to 0

while counter < 50:
    ##Accessing a random cell from columnA
    random_1 = sheet['A'+str(random.randint(1, lastRow))].value
    random_2 = sheet['B'+str(random.randint(1, lastRow))].value
    random_3 = sheet['C'+str(random.randint(1, lastRow))].value

    project_name = random_1 + ' ' + random_2 + ' ' + random_3.lower()
    print(project_name)
    counter += 1
